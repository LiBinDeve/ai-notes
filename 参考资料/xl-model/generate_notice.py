from openpyxl import Workbook
from openpyxl.styles import Side, Border, Font, Alignment


def _fill_notice_sheet(ws, month, name, data_rows, building='', room=''):
    """将催缴单内容写入给定的 worksheet，不创建/保存 workbook。"""
    font_title = Font(name='黑体', size=18, bold=True, color='000000')
    font_content = Font(name='仿宋', size=12, bold=False, color='000000')

    align_vertical = Alignment(horizontal='center', vertical='center', wrap_text=True)
    align_left = Alignment(horizontal='left', vertical='center', wrap_text=True)
    align_right = Alignment(horizontal='right', vertical='center', wrap_text=True)

    border_all = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    ws.merge_cells('A1:G1')
    ws.row_dimensions[1].height = 42
    ws['A1'].value = f"公寓 {month} 月租金缴费通知单"
    ws['A1'].font = font_title
    ws['A1'].alignment = align_vertical

    ws.merge_cells('A2:G2')
    ws.row_dimensions[2].height = 48
    ws['A2'].value = f"你好！你所承租的宏构公寓{building}{room}，在2026年{month}月应交各费用账单已产生，请核对后准时缴费，具体如下："
    ws['A2'].font = font_content
    ws['A2'].alignment = align_left

    for col, header in enumerate(['序号', '项目', '上月表底', '本月表底', '用量', '单价', '总价'], 1):
        cell = ws.cell(row=3, column=col)
        cell.value = header
        cell.font = font_content
        cell.alignment = align_vertical
        cell.border = border_all
    ws.row_dimensions[3].height = 24

    for row_num, data in data_rows:
        ws.row_dimensions[row_num].height = 24
        for col_num, value in enumerate(data, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = value
            cell.font = font_content
            cell.alignment = align_vertical
            cell.border = border_all

    ws.merge_cells('A9:G9')
    ws.row_dimensions[9].height = 24
    ws['A9'].value = f"备注：以上费用应于 2026 年 {month} 月 5 日前交清。"
    ws['A9'].font = font_content
    ws['A9'].alignment = align_left

    ws.merge_cells('A10:G10')
    ws.row_dimensions[10].height = 120
    ws['A10'].value = (
        "注：\n"
        "1.请以转账的形式交纳该费用，收费账户如下：\n"
        "户名：朱锦鹏\n"
        "账号:6227 0031 1221 0090 869\n"
        "开户行：中国建设银行 九江金信支行\n"
        "2.综合电费包含电梯用电及维护、公共照明、电损等费用。\n"
        "3.缴费后请保存回执或截图并发送到群上，以便提供证明。"
    )
    ws['A10'].font = font_content
    ws['A10'].alignment = align_left

    ws.merge_cells('A11:G11')
    ws.row_dimensions[11].height = 48
    ws['A11'].value = f"佛山宏构物业管理有限公司\n 2026 年 {month} 月 2 日"
    ws['A11'].font = font_content
    ws['A11'].alignment = align_right


def create_notice_excel(month, name, data_rows):
    wb = Workbook()
    del wb[wb.sheetnames[0]]
    ws = wb.create_sheet(title="催缴单")
    _fill_notice_sheet(ws, month, name, data_rows)
    wb.save(f"{month}_notice.xlsx")

