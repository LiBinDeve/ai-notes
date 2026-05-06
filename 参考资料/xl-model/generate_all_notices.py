from openpyxl import Workbook

from read_water_electricity import (
    read_water_electricity,
    TenantRecord,
    WaterRecord,
    ElectricityRecord,
)
from generate_notice import _fill_notice_sheet


def _v(x):
    """None 转空字符串，其余原样返回。"""
    return x if x is not None else ''


def build_data_rows(
    month: int,
    tenant: TenantRecord | None,
    water: WaterRecord,
    elec: ElectricityRecord,
) -> list:
    prev = month - 1 if month > 1 else 12
    rent = tenant.出租金额 if tenant else None

    return [
        (4, [1, f'{month}月租金',   '',               '',               '',               _v(rent),         _v(rent)]),
        (5, [2, '管理费',           '',               '',               '',               '',               '']),
        (6, [3, f'{month}月网费',   '',               '',               '',               '',               '']),
        (7, [4, f'{prev}月水费',    _v(water.上月读数), _v(water.本月读数), _v(water.实际用量), _v(water.单价),   _v(water.金额)]),
        (8, [5, f'{prev}月综合电费', _v(elec.上月读数),  _v(elec.本月读数),  _v(elec.实际用量),  _v(elec.单价),    _v(elec.总金额)]),
    ]


def generate_all_notices(
    month: int,
    source_file: str = "宏构公寓水电计算表.xlsx",
    output_path: str | None = None,
) -> str:
    """
    读取水电计算表，为每位有姓名的租户生成一张催缴单 sheet，
    全部写入同一个 Excel 文件。

    返回生成文件的路径。
    """
    tenants, water_list, elec_list = read_water_electricity(source_file)

    # 以租户姓名为键建立查找表
    tenant_by_name: dict[str, TenantRecord] = {
        t.租户姓名: t for t in tenants if t.租户姓名
    }
    elec_by_name: dict[str, ElectricityRecord] = {
        e.租户姓名: e for e in elec_list if e.租户姓名
    }

    wb = Workbook()
    del wb[wb.sheetnames[0]]

    seen_titles: set[str] = set()
    count = 0
    skipped: list[str] = []

    for water in water_list:
        name = water.租户姓名
        if not name:
            continue

        elec = elec_by_name.get(name)
        if elec is None:
            skipped.append(f"{name}（电表中无对应记录）")
            continue

        tenant = tenant_by_name.get(name)  # 租客基本情况可能尚未填写，允许为 None

        # 生成不重复的 sheet 标题（Excel 上限 31 字符）
        title = name[:31]
        if title in seen_titles:
            title = f"{name[:27]}_{count + 1}"[:31]
        seen_titles.add(title)

        ws = wb.create_sheet(title=title)
        data_rows = build_data_rows(month, tenant, water, elec)
        _fill_notice_sheet(ws, month, name, data_rows,
                           building=water.楼号 or '', room=water.房号 or '')
        count += 1

    if count == 0:
        print("没有找到任何有租户姓名的记录，文件未生成。")
        return ""

    path = output_path or f"{month}月催缴单.xlsx"
    wb.save(path)

    print(f"已生成 {count} 份催缴单 → {path}")
    if skipped:
        print(f"跳过 {len(skipped)} 条：{', '.join(skipped)}")

    return path


if __name__ == "__main__":
    generate_all_notices(month=5)
