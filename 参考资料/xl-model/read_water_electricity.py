import openpyxl
from dataclasses import dataclass
from typing import Optional
from datetime import date


EXCEL_FILE = "宏构公寓水电计算表.xlsx"


@dataclass
class TenantRecord:
    序号: Optional[int]
    楼号: Optional[str]
    楼层: Optional[str]
    房号: Optional[str]
    出租情况: Optional[str]
    租户姓名: Optional[str]
    身份证号: Optional[str]
    电话: Optional[str]
    备用电话: Optional[str]
    出租金额: Optional[float]
    押金: Optional[float]
    租期: Optional[str]
    起租日期: Optional[date]
    停租日期: Optional[date]
    是否续租: Optional[str]
    停租是否结清: Optional[str]


@dataclass
class WaterRecord:
    序号: Optional[int]
    租户姓名: Optional[str]
    楼号: Optional[str]
    楼层: Optional[str]
    房号: Optional[str]
    上月读数: Optional[float]
    本月读数: Optional[float]
    实际用量: Optional[float]
    单价: Optional[float]
    金额: Optional[float]


@dataclass
class ElectricityRecord:
    序号: Optional[int]
    租户姓名: Optional[str]
    楼号: Optional[str]
    楼层: Optional[str]
    房号: Optional[str]
    上月读数: Optional[float]
    本月读数: Optional[float]
    实际用量: Optional[float]
    单价: Optional[float]
    金额: Optional[float]
    电梯和公共用电量: Optional[float]
    公共金额: Optional[float]
    总金额: Optional[float]


def read_water_electricity(
    file_path: str = EXCEL_FILE,
) -> tuple[list[TenantRecord], list[WaterRecord], list[ElectricityRecord]]:
    """
    读取水电计算表，返回 (租客基本情况列表, 水表记录列表, 电表记录列表)。
    跳过表头行，None 值的行也会保留（对应空行/待填写行）。
    """
    wb = openpyxl.load_workbook(file_path, data_only=True)

    tenant_records: list[TenantRecord] = []
    for row in wb["租客基本情况"].iter_rows(min_row=2, values_only=True):
        tenant_records.append(TenantRecord(
            序号=row[0],
            楼号=row[1],
            楼层=row[2],
            房号=row[3],
            出租情况=row[4],
            租户姓名=row[5],
            身份证号=row[6],
            电话=row[7],
            备用电话=row[8],
            出租金额=row[9],
            押金=row[10],
            租期=row[11],
            起租日期=row[12],
            停租日期=row[13],
            是否续租=row[14],
            停租是否结清=row[15],
        ))

    water_records: list[WaterRecord] = []
    for row in wb["水"].iter_rows(min_row=2, values_only=True):
        water_records.append(WaterRecord(
            序号=row[0],
            租户姓名=row[1],
            楼号=row[2],
            楼层=row[3],
            房号=row[4],
            上月读数=row[5],
            本月读数=row[6],
            实际用量=row[7],
            单价=row[8],
            金额=row[9],
        ))

    electricity_records: list[ElectricityRecord] = []
    for row in wb["电"].iter_rows(min_row=2, values_only=True):
        electricity_records.append(ElectricityRecord(
            序号=row[0],
            租户姓名=row[1],
            楼号=row[2],
            楼层=row[3],
            房号=row[4],
            上月读数=row[5],
            本月读数=row[6],
            实际用量=row[7],
            单价=row[8],
            金额=row[9],
            电梯和公共用电量=row[10],
            公共金额=row[11],
            总金额=row[12],
        ))

    return tenant_records, water_records, electricity_records


if __name__ == "__main__":
    tenants, water, electricity = read_water_electricity()
    print(f"租客记录数: {len(tenants)}")
    print(f"水表记录数: {len(water)}")
    print(f"电表记录数: {len(electricity)}")

    print("\n--- 租客基本情况前3条 ---")
    for r in tenants[:3]:
        print(r)

    print("\n--- 水表前3条 ---")
    for r in water[:3]:
        print(r)

    print("\n--- 电表前3条 ---")
    for r in electricity[:3]:
        print(r)
